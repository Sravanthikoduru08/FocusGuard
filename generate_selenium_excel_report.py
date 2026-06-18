import json
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_report(json_path, browser_name, output_path):
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        sys.exit(1)
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selenium Test Report"
    
    # Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    normal_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
    
    status_styles = {
        "passed": {
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="375623")
        },
        "failed": {
            "fill": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="C00000")
        }
    }
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="1F4E78")
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # Title Block
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"FocusGuard E2E Selenium Test Results - {browser_name.upper()}"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Summary Info
    stats = data.get("stats", {})
    ws["A3"] = "Total Tests"
    ws["B3"] = stats.get("tests", 0)
    ws["A4"] = "Passed"
    ws["B4"] = stats.get("passes", 0)
    ws["A5"] = "Failed"
    ws["B5"] = stats.get("failures", 0)
    ws["A6"] = "Duration"
    ws["B6"] = f"{stats.get('duration', 0) / 1000:.2f}s"
    
    for row in range(3, 7):
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
        
    # Headers
    headers = ["Test Suite / Context", "Test Case Title", "Browser", "Duration", "Status", "Error Message"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    ws.row_dimensions[8].height = 25
    
    # Parse results
    rows_data = []
    results = data.get("results", [])
    for result in results:
        suites = result.get("suites", [])
        for suite in suites:
            suite_title = suite.get("title", "")
            tests = suite.get("tests", [])
            for test in tests:
                title = test.get("title", "")
                duration = f"{test.get('duration', 0) / 1000:.2f}s"
                state = test.get("state", "unknown")
                err_msg = ""
                err = test.get("err", {})
                if err and isinstance(err, dict):
                    err_msg = err.get("message", "")
                rows_data.append([suite_title, title, browser_name.capitalize(), duration, state, err_msg])
                
    # Write rows
    for r_idx, row_val in enumerate(rows_data, start=9):
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_val, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            if c_idx == 5: # Status
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val in status_styles:
                    cell.fill = status_styles[val]["fill"]
                    cell.font = status_styles[val]["font"]
            elif r_idx % 2 == 1:
                cell.fill = zebra_fill
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col_letter in ["A", "B", "C", "D", "E", "F"]:
            for cell in col:
                if cell.row in [1, 3, 4, 5, 6]:
                    continue
                val_str = str(cell.value or "")
                lines = val_str.split("\n")
                for line in lines:
                    max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws.views.sheetView[0].showGridLines = True
    wb.save(output_path)
    print(f"Excel report saved successfully to: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python generate_selenium_report.py <json_path> <browser_name> <output_path>")
        sys.exit(1)
    generate_report(sys.argv[1], sys.argv[2], sys.argv[3])
